try:
	import os
	import io
	from xlrd import open_workbook
	from xlutils.copy import copy
	from google_trans_new import google_translator
	from openpyxl import load_workbook
	from tqdm import tqdm
	import xlsxwriter
except:
	print('do pip install -r requirement.txt')

translator = google_translator()  

SHEETS= ['RCA','Raw Data']

def translate(dir_add, file_name, source, destination):
	if file_name.startswith('en_'):  return
	print('Working with '+file_name)
	rb = open_workbook(dir_add+file_name)
	wb = copy(rb)
	change=False
	for s in rb.sheet_names():
		if s in SHEETS:
			print('opening sheet '+s,end="\r")
			change=True
			rb_w = rb.sheet_by_name(s)
			wb_w = wb.get_sheet(s)
			for r in tqdm(range(rb_w.nrows)):
				to_do=''
				to_do_cell=[]
				for c, col in enumerate(rb_w.row_values(r)):
					if type(col)==str:
						masked_cell = col.replace('hide_data','@#@')
						masked_cell = masked_cell.replace('hide_data','@#@')
						to_do_cell.append((r, c))
						to_do= to_do + masked_cell + ' (*) '
				to_do = translator.translate(to_do,lang_tgt='en', lang_src=source)
				to_do = to_do.replace('( *)','(*)')
				to_do = to_do.replace('(* )','(*)')
				to_do = to_do.replace('( * )','(*)')
				to_do = to_do.replace('@ # @','@#@')
				to_do = to_do.replace('@ #@','@#@')
				to_do = to_do.replace('@# @','@#@')
				for (r, c), v in zip(to_do_cell, to_do.split('(*)')):
					v=v.replace('@#@',"hide_data")
					v=v.replace('(*)','')
					wb_w.write(r,c,v)
					
	
	if change: 
		wb.save('en_'+file_name[:-1])
		print('Converted to '+file_name[:-1])

def translatexo(dir_add, file_name, source, destination):
	if file_name.startswith('en_'):  return
	print('Working with '+file_name)
	with open(dir_add + file_name, "rb") as f:
		in_mem_file = io.BytesIO(f.read())
		wb = load_workbook(in_mem_file, data_only = True, keep_vba = True)
		change=False
		for s in wb.sheetnames:
			print('opening sheet '+s,end="\r")
			if s in SHEETS:
				change=True
				ws= wb[s]
				for row_cells in tqdm(ws.iter_rows()):
					to_do=''
					to_do_cell=[]
					for cell in row_cells:
						if cell.value and type(cell.value)==str:
							masked_cell = cell.value.replace('hide_data','@#@')
							masked_cell = masked_cell.replace('hide_data','@#@')
							# cell.value = str(translator.translate(masked_cell,lang_tgt='en', lang_src=source)).replace('@##@',"hide_data")
							to_do_cell.append(cell)
							to_do= to_do + masked_cell + ' (*) '
					# print(to_do)
					to_do = translator.translate(to_do,lang_tgt='en', lang_src=source)
					to_do = to_do.replace('( *)','(*)')
					to_do = to_do.replace('(* )','(*)')
					to_do = to_do.replace('( * )','(*)')
					to_do = to_do.replace('@ # @','@#@')
					to_do = to_do.replace('@ #@','@#@')
					to_do = to_do.replace('@# @','@#@')
					# print(to_do.split(' (*) '))
					for cell, v in zip(to_do_cell, to_do.split('(*)')):
						v=v.replace('@#@',"hide_data")
						v=v.replace('(*)','')
						cell.value=v
			else:
				wb.remove(wb.get_sheet_by_name(s))

		
		if change: 
			wb.save(dir_add+'en_'+file_name)
			print('Converted '+file_name)
			copy(open_workbook(dir_add+'en_'+file_name)).save(dir_add+'en_comatibility_mode'+file_name[:-1])
			print('Also Converted to comaptibilitymode')


def translatex(dir_add, file_name, source, destination):
	if file_name.startswith('en_'):  return
	print('Working with '+file_name)
	change=False
	wb = None
	rb = open_workbook(dir_add + file_name)
	for s in rb.sheet_names():
		if s in SHEETS:
			if not wb: wb = xlsxwriter.Workbook(dir_add+ 'en_'+file_name)
			wb.add_worksheet(s)
			print('opening sheet '+s,end="\r")
			change=True
			rb_w = rb.sheet_by_name(s)
			wb_w = wb.get_worksheet_by_name(s)
			for r in tqdm(range(rb_w.nrows)):
				for c, col in enumerate(rb_w.row_values(r)):
					if type(col)==str:
						masked_cell = col.replace('hide_data','@#$#@')
						masked_cell = masked_cell.replace('hide_data','@#$#@')
						translate_text = str(translator.translate(masked_cell,lang_tgt='en', lang_src=source)).replace('@#$#@',"hide_data")
						wb_w.write(r,c,translate_text)
					else:
						wb_w.write(r,c,col)
	if wb: 
		wb.close()
		print('saved to '+file_name)
	

def main():
	print('Select Mode \n 1. Try to keep formatting save in xls \n 2. Just Save in xlsx \n 3. Save in xls without formatting (fast)')
	mode=input('Enter 1 2 or 3\n')
	if mode=='1':
		print('Threr will be 2 traget files one without format in xls for comaptibility other one is xlsx with formatting in case xlsx is corrupt try another software or rename extension to xls')
	print('To add sheet names modify Constant in python file ')
	src=input('Enter source language code\n')
	dest=input('Enter destination language code\n')
	ext = input('Enter extension to crawl for in current directory and its subdir for translation. \n')
	to_work=[]
	for root, subdirs, files in os.walk(os.path.abspath('.')):
		for file in files:
			if file.endswith('.'+ext):
				if mode=='3':
					translate(root+'/', file, src, dest)
				elif mode=='2':
					translatex(root+'/', file, src, dest)
				elif mode=='1':
					translatexo(root+'/', file, src, dest)
				else:
					print('y'*10)

if __name__ == '__main__':
		main()