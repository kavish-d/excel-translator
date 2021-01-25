import io

import PySimpleGUI as sg
from google_trans_new import google_translator
from openpyxl import load_workbook
from xlrd import open_workbook
from xlutils.copy import copy

sg.theme("DarkTeal2")
translator = google_translator()
LANG_CODE = {'English': 'en', 'Spanish': 'es'}
fltr="afltr"

def get_sheet_names(faddr, fname, ext):
    if ext == 'xls':
        rb = open_workbook(faddr + fname)
        return rb.sheet_names()
    elif ext == 'xlsx':
        with open(faddr + fname, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
            wb = load_workbook(in_mem_file, data_only=True, keep_vba=True)
            return wb.sheetnames


def translate(dir_add, file_name, source, destination, sheets, window):
    if file_name.startswith('en_'):  return window
    rb = open_workbook(dir_add + file_name)
    wb = copy(rb)
    change = False
    window.close()
    for s in rb.sheet_names():
        if not sheets or s in sheets:
            change = True
            rb_w = rb.sheet_by_name(s)
            wb_w = wb.get_sheet(s)
            layout = [[sg.Text('Translating...' + s)],
                      [sg.ProgressBar(rb_w.nrows + 1, orientation='h', size=(20, 20), key='progressbar')]]
            window = sg.Window('Translate', layout, finalize=True)
            progress_bar = window['progressbar']
            for r in range(rb_w.nrows):
                progress_bar.Update(r + 1)
                to_do = ''
                to_do_cell = []
                for c, col in enumerate(rb_w.row_values(r)):
                    if type(col) == str:
                        masked_cell = col.replace(fltr, '@#@')
                        to_do_cell.append((r, c))
                        to_do = to_do + masked_cell + ' (*) '
                to_do = translator.translate(to_do, lang_tgt=destination, lang_src=source)
                to_do = to_do.replace('( *)', '(*)')
                to_do = to_do.replace('(* )', '(*)')
                to_do = to_do.replace('( * )', '(*)')
                to_do = to_do.replace('@ # @', '@#@')
                to_do = to_do.replace('@ #@', '@#@')
                to_do = to_do.replace('@# @', '@#@')
                for (r, c), v in zip(to_do_cell, to_do.split('(*)')):
                    v = v.replace('@#@', fltr)
                    v = v.replace('(*)', '')
                    wb_w.write(r, c, v)
            window.close()
    if change:
        wb.save('en_' + file_name)
        window.close()
        lay = [[sg.Text('Translatsed to en_' + file_name)]]
        window = sg.Window('Translator', lay)

    return window


def translatex(dir_add, file_name, source, destination, sheets, window):
    if file_name.startswith('en_'):  return window
    with open(dir_add + file_name, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
        wb = load_workbook(in_mem_file, data_only=True, keep_vba=True)
        change = False
        window.close()
        for s in wb.sheetnames:
            if not sheets or s in sheets:
                change = True
                ws = wb[s]
                layout = [[sg.Text('Translating...' + s)],
                          [sg.ProgressBar(ws.max_row + 1, orientation='h', size=(20, 20), key='progressbar')]]
                window = sg.Window('Translator', layout, finalize=True)
                progress_bar = window['progressbar']
                for i, row_cells in enumerate(ws.iter_rows()):
                    progress_bar.Update(i + 1)
                    to_do = ''
                    to_do_cell = []
                    for cell in row_cells:
                        if cell.value and type(cell.value) == str:
                            masked_cell = cell.value.replace(fltr, '@#@')
                            to_do_cell.append(cell)
                            to_do = to_do + masked_cell + ' (*) '
                    # print(to_do)
                    to_do = translator.translate(to_do, lang_tgt=destination, lang_src=source)
                    to_do = to_do.replace('( *)', '(*)')
                    to_do = to_do.replace('(* )', '(*)')
                    to_do = to_do.replace('( * )', '(*)')
                    to_do = to_do.replace('@ # @', '@#@')
                    to_do = to_do.replace('@ #@', '@#@')
                    to_do = to_do.replace('@# @', '@#@')
                    # print(to_do.split(' (*) '))
                    for cell, v in zip(to_do_cell, to_do.split('(*)')):
                        v = v.replace('@#@', fltr)
                        v = v.replace('(*)', '')
                        cell.value = v
                window.close()
            else:
                wb.remove(wb.get_sheet_by_name(s))

        if change:
            wb.save(dir_add + 'en_' + file_name[:-1])
            window.close()
            lay = [[sg.Text('Translated to en_' + file_name[:-1])]]
            window = sg.Window('Translator', lay)
            for i in range(500):
                pass
            copy(open_workbook(dir_add + 'en_' + file_name[:-1])).save(dir_add + 'en_comatibility_' + file_name[:-1])
            window.close()
            lay = [[sg.Text('Also created en_comatibility_' + file_name[:-1])]]
            window = sg.Window('Translator', lay)

    return window


faddr, fname, ext = '', '', ''
step = 0
layout = [[sg.T("")], [sg.Text("Choose a file: "), sg.Input(), sg.FileBrowse(key="-IN-")], [sg.Button("Load")]]
# layout = [[sg.Text("Hello from PySimpleGUI")], [sg.Button("OK")]]
window = sg.Window("Demo", layout)
# Create an event loop
while True:
    event, values = window.read()

    if event == sg.WIN_CLOSED or event == "Exit":
        break
    elif event == "Load":
        faddr = values["-IN-"]
        faddr = faddr.split('/' if '/' in faddr else '\\')
        fname, faddr = faddr[-1], '\\'.join(faddr[:-1] + [''])
        if fname.endswith('xls'):
            ext = 'xls'
        elif fname.endswith('xlsx'):
            ext = 'xlsx'
        print(faddr, fname, ext)
        sheets = get_sheet_names(faddr, fname, ext)
        sheet_selector = [[sg.Text("Select sheets to translate!")],
                          [sg.Listbox(values=['ALL'] + sheets, enable_events=True, size=(40, 20),
                                      select_mode=sg.LISTBOX_SELECT_MODE_MULTIPLE, key="-SHEETS-")],
                          [sg.Button("Submit")]
                          ]
        step += 1
        window.close()
        window = sg.Window("Demo", sheet_selector)
    elif event == "Submit":
        sheets = values["-SHEETS-"]
        if 'ALL' in sheets: sheets = []
        layout = [[sg.Text("Source Language"), sg.Combo(['English', 'Spanish'], enable_events=True, key='-LANG_SRC-')],
                  [sg.Text("Destination Language"),
                   sg.Combo(['English', 'Spanish'], enable_events=True, key='-LANG_DEST-')],
                  [sg.Button('Translate')]]
        step += 1
        window.close()
        window = sg.Window("Demo", layout)
    elif event == "Translate":
        src, dest = values['-LANG_SRC-'], values['-LANG_DEST-']
        print(src, dest)
        if src == dest:
            sg.Popup('Both are Same')
            continue
        else:
            src, dest = LANG_CODE[src], LANG_CODE[dest]
            if ext == 'xls':
                window = translate(faddr, fname, src, dest, sheets, window)
            elif ext == 'xlsx':
                window = translatex(faddr, fname, src, dest, sheets, window)
            break

window.close()
